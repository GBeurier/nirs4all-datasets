"""Discover datasets in a source tree and auto-generate schema-valid descriptors.

Walks a folder tree of NIRS datasets (the nirs4all ``Xtrain/Ytrain/...`` convention, organized as
``<task>/<family>/<leaf>``) and writes one ``catalog/datasets/<id>.yaml`` per leaf, enriched from an
optional ``DatabaseDetail.xlsx`` master sheet. Descriptors are marked ``generation.managed = true`` so
re-running only overwrites machine-generated ones (never a human-edited descriptor) and only when the
processing-relevant content changed (``descriptor_hash``).

Boundary: file→role mapping reuses nirs4all's ``FolderParser`` (no naming rules re-implemented here).
Everything else is descriptor *authoring* — axis-unit inference, license→SPDX mapping, honest
governance defaults — which is this package's own domain. No NIRS/ML logic is re-implemented.

Local-only stance: descriptors default to ``visibility: restricted`` (workflow state = "not published
yet"; the Dataverse collections are closed) with an honest ``confidentiality_class`` (``public`` for an
open license, else ``internal``). Publishing later is a deliberate, separate action.
"""
from __future__ import annotations

import hashlib
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import pandas as pd
import yaml

from nirs4all_datasets.manifest import metadata_hash
from nirs4all_datasets.schema import (
    _OPEN_LICENSES,
    AxisUnit,
    ConfidentialityClass,
    DataCite,
    DatasetDescriptor,
    Generation,
    Governance,
    Instrument,
    Modality,
    OriginSource,
    Provenance,
    PublicationRef,
    SignalType,
    SourceAccess,
    SourceKind,
    SourceMode,
    Target,
    TaskType,
    Visibility,
)

GENERATOR = "discover"
GENERATOR_VERSION = "2"

# Trees under a NIRS DB source root that hold reference-ready leaves. Everything else
# (``Publications/`` PDFs, ``chantiers/`` and ``v2.0/`` work-in-progress, ``unusableDB/``) is skipped.
_LEAF_TREES = ("regression", "classification", "multimachines")

# Data-repository DOI prefixes -> (OriginSource.kind, human repository name). A DOI under one of these
# is where the *bytes* live (-> sources[]); any other 10.x DOI is a journal paper (-> citation).
_DATA_DOI_PREFIXES: dict[str, tuple[str, str]] = {
    "10.57745": ("dataverse", "Recherche Data Gouv"),
    "10.15454": ("dataverse", "Portail Data INRAE"),
    "10.18710": ("dataverse", "DataverseNO"),
    "10.7910": ("dataverse", "Harvard Dataverse"),
    "10.34725": ("dataverse", "Dataverse"),
    "10.18167": ("dataverse", "CIRAD Dataverse"),
    "10.5281": ("zenodo", "Zenodo"),
    "10.6084": ("figshare", "figshare"),
    "10.5061": ("url", "Dryad"),
}
_DOI_IN_TEXT = re.compile(r"(10\.\d{4,9}/[^\s,;\"'<>]+)")
# Target names that are categorical/identifier regardless of dtype (an integer-encoded ``*_type`` or a
# numeric ``row_no`` is still not a regression target). Applied to a token-normalized name (camelCase +
# separators -> ``_``); whole-token match so soil props like ``rubidium``/``carbon`` don't false-hit.
_CLF_NAME_RE = re.compile(
    r"(?:^|_)(id|type|class|category|categorical|label|group|species|genus|family|name|code|smiles|inchi|inchikey|formula|variety|cultivar|origin|authenticity|adulteration|material|mineral|subclass|date|datetime|timestamp|row|rowno|unnamed|observation)(?:_|$)",
    re.IGNORECASE,
)
_NA_TOKENS = frozenset({"", "nan", "na", "null", "none", "n/a", "<na>", "."})


def _norm_key(name: str) -> str:
    """Normalize a column name for robust matching (lowercase, drop all non-alphanumerics)."""
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def _norm_tokens(name: str) -> str:
    """Token form of a name: split camelCase, collapse separators to ``_`` (so ``LeafNumber`` -> ``leaf_number``, ``plant.id`` -> ``plant_id``)."""
    spaced = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", str(name))
    return re.sub(r"[^a-z0-9]+", "_", spaced.lower()).strip("_")
# Instrument tokens encoded in multimachines leaf names (e.g. ``An_..._byCultivar_MicroNIR_NeoSpectra``).
_INSTRUMENT_TOKENS = (("micronir_neospectra", "MicroNIR + NeoSpectra"), ("micronir", "MicroNIR"), ("neospectra", "NeoSpectra"), ("asd", "ASD"))


def _extract_doi(text: str) -> str | None:
    """Pull a bare DOI out of a ``doi:`` / ``https://doi.org/...`` / raw string; None if absent."""
    match = _DOI_IN_TEXT.search(str(text))
    return match.group(1).rstrip(".") if match else None


def _classify_origin(raw: Any) -> tuple[str, dict[str, Any] | str | None]:
    """Classify a master-sheet ``Source``/``Ref`` cell.

    Returns ``('source', origin_kwargs)`` for a data home (DOI in a data repo, or a data URL),
    ``('paper', doi)`` for a journal/publisher DOI, ``('manual', text)`` for a non-DOI non-URL note,
    or ``('skip', None)`` when blank. This is what keeps a *paper* DOI out of ``sources[]``.
    """
    if raw is None:
        return ("skip", None)
    text = str(raw).strip()
    if not text or text.lower() in ("none", "nan", "n/a"):
        return ("skip", None)
    doi = _extract_doi(text)
    if doi:
        prefix = doi.split("/", 1)[0]
        if prefix in _DATA_DOI_PREFIXES:
            kind, repo = _DATA_DOI_PREFIXES[prefix]
            return ("source", {"kind": kind, "mode": "raw", "locator": doi, "access": "open", "title": repo})
        return ("paper", doi)  # any other DOI is a journal/publisher DOI -> citation, not a data source
    if text.lower().startswith("http"):
        host = urlparse(text).netloc.lower()
        if "zenodo.org" in host:
            return ("source", {"kind": "zenodo", "mode": "raw", "locator": text, "access": "open", "title": "Zenodo"})
        if "github.com" in host:
            return ("source", {"kind": "url", "mode": "raw", "locator": text, "access": "open", "title": "GitHub"})
        # registration/click-through portals must be fetched by hand.
        access = "manual" if any(h in host for h in ("esdac.jrc", "eigenvector.com", "ucphchemometrics", "rdrr.io")) else "open"
        return ("source", {"kind": "url", "mode": "raw", "locator": text, "access": access})
    return ("manual", text)


def _instrument_model(name: str) -> str | None:
    """Device encoded in a (multimachines) leaf name, e.g. ``..._MicroNIR_NeoSpectra`` -> that label."""
    low = name.lower()
    for token, label in _INSTRUMENT_TOKENS:
        if token in low:
            return label
    return None

# Y headers too generic to use as a target name (fall back to the xlsx trait / leaf name). Compared
# after normalizing (lower-case, alphanumerics only), so "y_cal"/"X.1" match. "x"/"v1" are R's
# default column names for an unnamed exported vector.
_GENERIC_TARGET = frozenset({"y", "ycal", "yval", "ytrain", "ytest", "target", "value", "label", "class", "labels", "x", "v1", "x1", "0", "unnamed0"})


def _is_generic_header(header: str | None) -> bool:
    if not header:
        return False
    return re.sub(r"[^a-z0-9]", "", header.lower()) in _GENERIC_TARGET

# Master-sheet licence strings → SPDX identifiers (lower-cased keys).
_SPDX = {
    "cc0 1.0": "CC0-1.0",
    "cc-by-4.0": "CC-BY-4.0",
    "cc by 4.0": "CC-BY-4.0",
    "creative commons attribution 4.0 international": "CC-BY-4.0",
    "cc-by-sa-4.0": "CC-BY-SA-4.0",
    "etalab 2.0": "etalab-2.0",
    "gpl (>= 2)": "GPL-2.0-or-later",
    "mit license": "MIT",
    "mit": "MIT",
    "odbl-1.0": "ODbL-1.0",
}


def slugify(text: str) -> str:
    """Lowercase slug matching the descriptor id pattern ``^[a-z0-9]+(_[a-z0-9]+)*$``."""
    slug = re.sub(r"[^a-z0-9]+", "_", str(text).lower()).strip("_")
    return re.sub(r"_+", "_", slug)


def dataset_id(family: str, name: str) -> str:
    """Stable id ``<family>_<leaf>`` (de-duplicating the family prefix when the leaf already carries it)."""
    fam, leaf = slugify(family), slugify(name)
    base = leaf if (fam and (leaf == fam or leaf.startswith(fam + "_"))) else f"{fam}_{leaf}" if fam else leaf
    return base or "dataset"


def _folder_config(path: Path) -> dict[str, Any]:
    """nirs4all's parsed folder config (file→role), or ``{}`` if not a dataset folder."""
    from nirs4all.data.parsers.folder_parser import FolderParser

    result = FolderParser().parse(str(path))
    if not result.success:
        return {}
    return {k: v for k, v in result.config.items() if v is not None}


@dataclass
class Leaf:
    """One discovered dataset directory."""

    path: Path
    source_relpath: str
    family: str
    name: str
    task_root: str  # "regression" | "classification"
    config: dict[str, Any]


def find_leaves(source_root: str | Path) -> list[Leaf]:
    """Find every reference-ready leaf dataset under ``<source_root>/{regression,classification,multimachines}``.

    A leaf is a directory nirs4all's ``FolderParser`` maps to a training X file (so family folders that
    only hold sub-datasets, papers, or raw extras are skipped). ``multimachines`` leaves are treated as
    regression; any path with a ``_todo``/``todo`` segment is skipped (work-in-progress).
    """
    root = Path(source_root)
    leaves: list[Leaf] = []
    for tree in _LEAF_TREES:
        base = root / tree
        if not base.is_dir():
            continue
        task_root = "classification" if tree == "classification" else "regression"
        for d in sorted(p for p in base.rglob("*") if p.is_dir()):
            rel = d.relative_to(base)
            if any(part.lower() == "todo" or part.lower().endswith("_todo") for part in rel.parts):
                continue  # work-in-progress leaf/family
            config = _folder_config(d)
            if not config.get("train_x"):
                continue
            family = rel.parts[0]
            leaves.append(Leaf(path=d, source_relpath=d.relative_to(root).as_posix(), family=family, name=d.name, task_root=task_root, config=config))
    return leaves


def _first_path(value: Any) -> Path | None:
    """The first path from a config entry that may be a single path or a multi-source list."""
    if isinstance(value, list):
        return Path(value[0]) if value else None
    return Path(value) if isinstance(value, str) else None


def _infer_axis_unit(x_path: Path) -> str:
    """Infer the spectral axis unit from the X header row (verified rule for this corpus).

    Strips quotes and ``_nm``/``cm-1`` suffixes, then: explicit suffix wins; otherwise numeric headers
    are ``nm`` when their max ≤ 2600 (classic VIS/NIR in nm) and ``cm-1`` above (FT-NIR/Raman in cm⁻¹);
    non-numeric headers fall back to ``index``.
    """
    try:
        line = x_path.open("r", encoding="utf-8", errors="replace").readline().strip()
    except OSError:
        return AxisUnit.NONE.value
    delim = ";" if line.count(";") >= line.count(",") else ","
    tokens = [t.strip().strip('"').strip("'") for t in line.split(delim) if t.strip()]
    sample = " ".join(tokens[:5]).lower()
    if "nm" in sample:
        return AxisUnit.WAVELENGTH.value
    if "cm-1" in sample or "cm_1" in sample:
        return AxisUnit.WAVENUMBER.value
    values: list[float] = []
    for tok in tokens:
        cleaned = re.sub(r"[^0-9.+\-eE]", "", tok.replace(",", "."))
        try:
            values.append(float(cleaned))
        except ValueError:
            continue
    if len(values) < max(2, 0.5 * len(tokens)):
        return AxisUnit.INDEX.value
    return AxisUnit.WAVELENGTH.value if max(values) <= 2600 else AxisUnit.WAVENUMBER.value


def _read_column(path: Path | None) -> tuple[str | None, pd.Series | None]:
    """Read the first column (header + values) of a Y file; ``(None, None)`` on failure.

    Delimiter-aware (``;``/``,``) rather than sniffer-based: a single-column header like ``Ycal`` or
    ``CoffeeType`` has no delimiter, so the sniffer would wrongly split it on an interior character.
    """
    if path is None or not path.exists() or path.stat().st_size == 0:
        return None, None
    try:
        lines = [ln for ln in path.read_text(encoding="utf-8", errors="replace").splitlines() if ln.strip()]
    except OSError:
        return None, None
    if not lines:
        return None, None
    delim = ";" if lines[0].count(";") >= lines[0].count(",") else ","

    def _first(line: str) -> str:
        return line.split(delim)[0].strip().strip('"').strip("'")

    header = _first(lines[0])
    return (header or None), pd.Series([_first(ln) for ln in lines[1:]], dtype="object")


def _trait_from_leaf(family: str, name: str) -> str | None:
    """First leaf-name token after the family prefix (a decent trait guess, e.g. Corn_**Oil** → ``oil``)."""
    fam, leaf = slugify(family), slugify(name)
    if fam and leaf.startswith(fam + "_"):
        leaf = leaf[len(fam) + 1 :]
    token = leaf.split("_")[0] if leaf else ""
    return token or None


def _target_info(leaf: Leaf) -> tuple[str | None, list[str] | None, str | None]:
    """Return ``(y_header, class_names_or_None, warning_or_None)`` for the leaf's target.

    For classification, class names are the sorted unique string labels over train **and** test (an
    honest set; the canonical Y stores nirs4all-assigned integer indices, whose exact index↔name
    mapping is intentionally not asserted here).
    """
    header, train_col = _read_column(_first_path(leaf.config.get("train_y")))
    if leaf.task_root != "classification":
        return header, None, None
    _, test_col = _read_column(_first_path(leaf.config.get("test_y")))
    labels: set[str] = set()
    for col in (train_col, test_col):
        if col is not None:
            labels |= {str(v).strip() for v in col.tolist() if str(v).strip() and str(v).lower() != "nan"}
    classes = sorted(labels)
    warning = None if classes else "could not read class labels from Y files"
    return header, (classes or None), warning


def load_xlsx(xlsx_path: str | Path) -> dict[str, dict[str, Any]]:
    """Index the ``DatabaseDetail.xlsx`` master sheet by the (lower-cased) ``Dataset`` leaf name.

    Also adds ``family::<database>`` fallback keys (the first row seen per family) so split-variant
    leaves that are not themselves a sheet row can still inherit the family-constant provenance
    (``Source``/``Ref``/``Licence``).
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    index: dict[str, dict[str, Any]] = {}
    family: dict[str, dict[str, Any]] = {}
    for i, row in enumerate(rows[1:], start=2):
        record = {header[j]: row[j] for j in range(min(len(header), len(row)))}
        record["_row"] = i
        key = record.get("dataset")
        if key is not None and str(key).strip():
            index[str(key).strip().lower()] = record
        fam = record.get("database")
        if fam is not None and str(fam).strip():
            family.setdefault(str(fam).strip().lower(), record)
    for fam, record in family.items():
        index.setdefault(f"family::{fam}", record)
    return index


def _spdx(license_str: Any) -> str | None:
    """Map a master-sheet licence string to an SPDX id / ``LicenseRef-*``; ``None`` when unknown/blank."""
    if license_str is None:
        return None
    key = str(license_str).strip().lower()
    if not key or key in ("none", "nan", "n/a"):
        return None
    if key in _SPDX:
        return _SPDX[key]
    if "lucas" in key:
        return "LicenseRef-LUCAS-SOIL"
    return f"LicenseRef-{slugify(license_str)}" if slugify(license_str) else None


def _governance(license_spdx: str | None, owner: str) -> Governance:
    """Honest governance for a local (unpublished) descriptor.

    ``confidentiality_class`` is factual (``public`` for an open licence, else ``internal``);
    ``visibility`` stays ``restricted`` — the workflow state "not published yet", not a confidentiality
    claim. Open-data fields are filled so the descriptor could be published later after review.
    """
    is_open = license_spdx in _OPEN_LICENSES
    return Governance(
        license=license_spdx or "LicenseRef-unknown",
        visibility=Visibility.RESTRICTED,
        confidentiality_class=ConfidentialityClass.PUBLIC if is_open else ConfidentialityClass.INTERNAL,
        owner_steward=owner,
        redistribution_rights=(f"Open redistribution under {license_spdx}." if is_open else "Original source licence applies; verify before redistribution."),
        consent_ethics_status="Not assessed (auto-generated; verify before release).",
        anonymization_status="Not assessed (auto-generated; verify before release).",
        permitted_use="Research and benchmarking.",
        access_policy="Local use; not published to Dataverse.",
    )


def _leaf_fingerprint(leaf: Leaf) -> str:
    """Cheap content fingerprint: sha256 over the leaf's sorted ``(name, size)`` data files."""
    parts = [f"{p.name}:{p.stat().st_size}" for p in sorted(leaf.path.iterdir()) if p.is_file()]
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


def build_descriptor(leaf: Leaf, xlsx_index: dict[str, dict[str, Any]]) -> tuple[DatasetDescriptor, list[str]]:
    """Build a schema-valid descriptor for one leaf, enriched from the master sheet where matched."""
    warnings: list[str] = []
    meta = xlsx_index.get(leaf.name.lower(), {})
    fam_meta = xlsx_index.get(f"family::{leaf.family.lower()}", {})

    def _meta(key: str) -> Any:
        """Leaf row value, falling back to the family row for family-constant provenance."""
        value = meta.get(key)
        return value if value not in (None, "") else fam_meta.get(key)

    did = dataset_id(leaf.family, leaf.name)

    y_header, classes, target_warn = _target_info(leaf)
    if target_warn:
        warnings.append(target_warn)
    if leaf.task_root == "classification":
        task_type = "binary_classification" if (classes and len(classes) == 2) else "multiclass_classification"
    else:
        task_type = "regression"

    trait = str(meta["trait"]).strip() if meta.get("trait") else None
    if y_header and not _is_generic_header(y_header):
        target_name = y_header
    elif trait:
        target_name = trait
    else:
        target_name = _trait_from_leaf(leaf.family, leaf.name) or slugify(leaf.name) or "target"

    axis_unit = _infer_axis_unit(_first_path(leaf.config.get("train_x")) or leaf.path)
    license_spdx = _spdx(_meta("licence") or _meta("license"))

    # Classify the master-sheet Source (data home) and Ref (paper): data DOIs/URLs -> sources[];
    # journal DOIs -> related_publications. The owner/steward is the hosting repository (an honest
    # steward), never the raw DOI string that the old generator stuffed into contributor.
    sources: list[OriginSource] = []
    publications: list[PublicationRef] = []
    steward: str | None = None
    seen_loc: set[str] = set()
    seen_doi: set[str] = set()
    for cell in (_meta("source"), _meta("ref")):
        kind, payload = _classify_origin(cell)
        if kind == "source" and isinstance(payload, dict):
            if payload["locator"] in seen_loc:
                continue
            seen_loc.add(payload["locator"])
            try:
                sources.append(OriginSource(**payload))
                steward = steward or payload.get("title")
            except Exception as exc:  # noqa: BLE001 - a malformed cell must not abort the leaf
                warnings.append(f"unparseable origin source {cell!r}: {exc}")
        elif kind == "paper" and isinstance(payload, str) and payload not in seen_doi:
            seen_doi.add(payload)
            publications.append(PublicationRef(doi=payload))
        elif kind == "manual" and isinstance(payload, str) and payload not in seen_loc:
            seen_loc.add(payload)
            sources.append(OriginSource(kind=SourceKind.MANUAL, locator=payload, access=SourceAccess.MANUAL))
    owner = steward or "NIRS DB reference collection"

    sample = str(meta["sample"]).strip() if meta.get("sample") else None
    split = str(meta["split"]).strip() if meta.get("split") else None
    descr_bits = [f"{leaf.family} dataset ({leaf.task_root})."]
    if sample:
        descr_bits.append(f"Sample: {sample}.")
    if trait:
        descr_bits.append(f"Trait: {trait}.")
    if split:
        descr_bits.append(f"Split: {split}.")
    descr_bits.append("Auto-generated descriptor (verify before publication).")

    keywords = [k for k in [slugify(leaf.family), "nir", leaf.task_root, slugify(trait) if trait else None] if k]
    citation = f"https://doi.org/{publications[0].doi}" if publications else None

    descriptor = DatasetDescriptor(
        id=did,
        name=f"{leaf.family} — {leaf.name}",
        version="0.1.0",
        description=" ".join(descr_bits),
        domain=leaf.family.lower(),
        keywords=keywords,
        citation=citation,
        instrument=Instrument(modality=Modality.NIR, model=_instrument_model(leaf.name), axis_unit=AxisUnit(axis_unit), signal_type=SignalType.AUTO),
        targets=[Target(name=target_name, task_type=TaskType(task_type), unit=None, classes=classes)],
        provenance=Provenance(
            contributor=owner,
            reference_method=trait,
            ingest_reader="tabular",
            known_exclusions=None,
        ),
        governance=_governance(license_spdx, owner),
        datacite=DataCite(related_publications=publications) if publications else None,
        sources=sources,
        generation=Generation(
            managed=True,
            generator=GENERATOR,
            generator_version=GENERATOR_VERSION,
            source_relpath=leaf.source_relpath,
            source_fingerprint=_leaf_fingerprint(leaf),
            xlsx_row=int(meta["_row"]) if meta.get("_row") else None,
        ),
    )
    return descriptor, warnings


GENERATOR_V2 = "discover-v2"


def _dir_fingerprint(path: Path) -> str:
    """Cheap content fingerprint of a leaf dir: sha256 over its sorted ``(name, size)`` files."""
    parts = [f"{p.name}:{p.stat().st_size}" for p in sorted(path.iterdir()) if p.is_file()]
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


def _v2_axis_unit(blocks: list[dict[str, Any]]) -> str:
    """Map a v2.0 spectral block's ``axis_unit`` onto our enum (only nm/cm-1 are representable)."""
    unit = str((blocks[0].get("axis_unit") if blocks else "") or "").strip().lower()
    if "cm-1" in unit or "wavenumber" in unit:
        return AxisUnit.WAVENUMBER.value
    if unit in ("nm", "nanometer", "nanometers") or "wavelength (nm)" in unit:
        return AxisUnit.WAVELENGTH.value
    return AxisUnit.NONE.value  # micrometers / unknown -> not representable; declared honestly as none


def _v2_numeric_columns(y_path: Path, sample_rows: int = 80) -> dict[str, bool]:
    """Per-column numeric-ness of a v2.0 ``Y.csv`` (header + a sample), to type targets without an
    explicit ``target_types`` entry: mostly-numeric -> regression, else classification."""
    if not y_path.exists():
        return {}
    lines = [ln for ln in y_path.read_text(encoding="utf-8", errors="replace").splitlines() if ln.strip()][: sample_rows + 1]
    if len(lines) < 2:
        return {}
    delim = ";" if lines[0].count(";") >= lines[0].count(",") else ","
    headers = [h.strip().strip('"').strip("'") for h in lines[0].split(delim)]
    numeric: dict[str, bool] = {}
    for i, header in enumerate(headers):
        vals = [row.split(delim)[i].strip().strip('"').strip("'") for row in lines[1:] if i < len(row.split(delim))]
        nonempty = [v for v in vals if v.lower() not in _NA_TOKENS]
        if not nonempty:
            numeric[_norm_key(header)] = True  # no evidence -> default regression
            continue
        ok = sum(1 for v in nonempty if _is_float(v))
        numeric[_norm_key(header)] = ok >= 0.8 * len(nonempty)
    return numeric


def _is_float(value: str) -> bool:
    try:
        float(value.replace(",", "."))
        return True
    except ValueError:
        return False


def find_v2_leaves(source_root: str | Path) -> list[Path]:
    """v2.0 standardized leaves: dirs under ``v2.0/`` holding a ``dataset_card.json`` + ``Y.csv``.

    These are Frictionless-style packages (machine-readable card declaring targets/sources/license), so
    the descriptor is authored from the card -- no free-text guessing.
    """
    base = Path(source_root) / "v2.0"
    if not base.is_dir():
        return []
    return sorted(d for d in base.iterdir() if d.is_dir() and (d / "dataset_card.json").exists() and (d / "Y.csv").exists())


def build_v2_descriptor(leaf_dir: Path) -> tuple[DatasetDescriptor, list[str]]:
    """Author a descriptor for one v2.0 package from its authoritative ``dataset_card.json``.

    Targets come from ``target_summary.target_variables`` (so metadata columns are never mistaken for
    targets); sources from ``detected_sources`` (+ the maintainer ``source_to_standard.py`` script);
    governance from ``license_summary`` (``public_release_allowed: false`` -> restricted/internal).
    """
    warnings: list[str] = []
    card = json.loads((leaf_dir / "dataset_card.json").read_text(encoding="utf-8"))
    did = slugify(card.get("dataset_id") or leaf_dir.name)

    tsum = card.get("target_summary") or {}
    tvars = [str(v) for v in (tsum.get("target_variables") or [])]
    ttypes = {str(k): str(v).lower() for k, v in (tsum.get("target_types") or {}).items()}
    y_numeric = _v2_numeric_columns(leaf_dir / "Y.csv")

    def _task(var: str) -> TaskType:
        # A categorical/identifier *name* (tablet_type, sample_id, smiles, molecular_formula, Row_No,
        # Data_Collection_Date, ...) is never a regression target -- this overrides even an explicit
        # card type, which is demonstrably wrong for these. Matched on a token-normalized name.
        if _CLF_NAME_RE.search(_norm_tokens(var)):
            return TaskType.MULTICLASS_CLASSIFICATION
        kind = ttypes.get(var, "")
        if kind.startswith("regress"):
            return TaskType.REGRESSION
        if kind and ("class" in kind or "label" in kind or "categor" in kind):
            return TaskType.MULTICLASS_CLASSIFICATION
        # No explicit type (e.g. OSSL's many soil-chemistry targets): infer from the Y column dtype
        # (normalized name match), defaulting to regression (the NIRS norm) -- never blanket-classification.
        return TaskType.MULTICLASS_CLASSIFICATION if y_numeric.get(_norm_key(var)) is False else TaskType.REGRESSION

    targets = [Target(name=v, task_type=_task(v)) for v in tvars]
    if not targets:
        targets = [Target(name="target", task_type=TaskType.REGRESSION)]
        warnings.append("dataset_card.json declares no target_variables; placeholder target used")

    # Harvest origin URLs/DOIs from everywhere the card records them (detected_sources, source_summary,
    # official_source); ignore local filesystem paths. The standardization script is a maintainer source.
    sources: list[OriginSource] = []
    publications: list[PublicationRef] = []
    seen_loc: set[str] = set()
    seen_doi: set[str] = set()
    candidates: list[Any] = [e.get("url") for e in (card.get("detected_sources") or []) if e.get("url")]
    candidates.append((card.get("source_summary") or {}).get("source_url"))
    official = card.get("official_source") or {}
    for key in ("download_urls", "landing_urls", "urls", "download_page", "landing_page", "source_page", "url"):
        value = official.get(key)
        candidates.extend(value if isinstance(value, list) else [value])
    for raw_url in candidates:
        if not raw_url or not (str(raw_url).startswith("http") or _extract_doi(str(raw_url))):
            continue  # skip blanks + local filesystem paths (e.g. D:\... cache paths)
        kind, payload = _classify_origin(raw_url)
        if kind == "source" and isinstance(payload, dict) and payload["locator"] not in seen_loc:
            seen_loc.add(payload["locator"])
            try:
                sources.append(OriginSource(**payload))
            except Exception as exc:  # noqa: BLE001 - a malformed url must not abort the leaf
                warnings.append(f"unparseable v2 source {raw_url!r}: {exc}")
        elif kind == "paper" and isinstance(payload, str) and payload not in seen_doi:
            seen_doi.add(payload)
            publications.append(PublicationRef(doi=payload))
    if (leaf_dir / "source_to_standard.py").exists():
        sources.append(OriginSource(kind=SourceKind.SCRIPT, mode=SourceMode.RAW, locator="source_to_standard.py", access=SourceAccess.MANUAL, title="standardization script (maintainer-only)"))

    lic = card.get("license_summary") or {}
    rights = card.get("rights") or {}
    public_ok = bool(lic.get("public_release_allowed", rights.get("public_release_allowed", False)))
    source_name = str((card.get("source_summary") or {}).get("source_name") or "NIRS DB v2.0 collection")
    governance = Governance(
        license="LicenseRef-not-cleared" if not public_ok else "LicenseRef-review-before-use",
        visibility=Visibility.RESTRICTED,
        confidentiality_class=ConfidentialityClass.INTERNAL if not public_ok else ConfidentialityClass.PUBLIC,
        owner_steward=source_name,
        redistribution_rights=str(lic.get("rights_notes") or "Redistribution not cleared; verify source terms before release."),
        consent_ethics_status="Not assessed (auto-generated from v2.0 package; verify before release).",
        anonymization_status="Not assessed (auto-generated from v2.0 package; verify before release).",
        permitted_use="Research and benchmarking; private use only." if not public_ok else "Research and benchmarking.",
        access_policy="Local use; not published to Dataverse." + ("" if public_ok else " Manual download / private-use-only per source."),
    )

    blocks = card.get("spectral_blocks") or []
    model = str(blocks[0].get("instrument_name")) if blocks and blocks[0].get("instrument_name") else None
    instrument = Instrument(modality=Modality.NIR, model=model, axis_unit=AxisUnit(_v2_axis_unit(blocks)), signal_type=SignalType.AUTO)
    n_sources = max(1, len(blocks))
    family = slugify(did.split("_")[0]) or None

    descriptor = DatasetDescriptor(
        id=did,
        name=str(card.get("dataset_name") or leaf_dir.name),
        version="0.1.0",
        description=f"{card.get('dataset_name') or did}. v2.0 standardized package: {len(tvars)} target(s), {n_sources} spectral block(s). Auto-generated from dataset_card.json (verify before publication).",
        domain=family,
        keywords=[k for k in ["nir", "v2", family] if k],
        citation=f"https://doi.org/{publications[0].doi}" if publications else None,
        instrument=instrument,
        targets=targets,
        n_sources=n_sources,
        provenance=Provenance(contributor=source_name, ingest_reader="tabular", warnings=warnings),
        governance=governance,
        datacite=DataCite(related_publications=publications) if publications else None,
        sources=sources,
        generation=Generation(
            managed=True,
            generator=GENERATOR_V2,
            generator_version=GENERATOR_VERSION,
            source_relpath=f"v2.0/{leaf_dir.name}",
            source_fingerprint=_dir_fingerprint(leaf_dir),
        ),
    )
    return descriptor, warnings


def _write_descriptor(descriptor: DatasetDescriptor, path: Path) -> None:
    """Write a descriptor as clean YAML with an auto-generated banner."""
    data = descriptor.model_dump(mode="json", exclude_none=True)
    banner = "# AUTO-GENERATED by nirs4all-datasets `discover`. Edit freely; set generation.managed: false to protect from regeneration.\n"
    path.write_text(banner + yaml.safe_dump(data, sort_keys=False, allow_unicode=True), encoding="utf-8")


def bootstrap(source_root: str | Path, catalog_root: str | Path, *, xlsx_path: str | Path | None = None, force: bool = False, prune: bool = False) -> dict[str, Any]:
    """Generate/refresh descriptors for every leaf under ``source_root``.

    Idempotent: a human-edited descriptor (``generation.managed`` false/absent) is never overwritten;
    a managed one is rewritten only when ``force`` or its ``metadata_hash`` changed (so a provenance-only
    edit, e.g. ``sources``/``citation`` — excluded from ``descriptor_hash`` — still triggers a refresh).

    Reconciliation (``new replaces old``): a *managed* descriptor whose id the source no longer
    produces is an orphan. With ``prune`` it is deleted (human-authored orphans are always kept and
    only flagged). A committed ``catalog/reconciliation.json`` records the full add/update/remove diff.
    Returns a report ``{created, updated, skipped, errors, ids, removed, kept_human_orphans, pruned}``.
    """
    out_dir = Path(catalog_root) / "catalog" / "datasets"
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_index = load_xlsx(xlsx_path) if xlsx_path and Path(xlsx_path).exists() else {}

    report: dict[str, Any] = {"created": [], "updated": [], "skipped": [], "errors": [], "ids": {}}

    # Build descriptors. v2.0 standardized packages are emitted FIRST so they win id collisions over the
    # raw regression/ equivalents (v2.0 is the cleaner, machine-described version).
    items: list[tuple[DatasetDescriptor, list[str], str]] = []
    for v2dir in find_v2_leaves(source_root):
        rel = f"v2.0/{v2dir.name}"
        try:
            descriptor, warns = build_v2_descriptor(v2dir)
            items.append((descriptor, warns, rel))
        except Exception as exc:  # noqa: BLE001 - one bad package must not abort the sweep
            report["errors"].append({"leaf": rel, "error": f"{type(exc).__name__}: {exc}"})
    for leaf in find_leaves(source_root):
        try:
            descriptor, warns = build_descriptor(leaf, xlsx_index)
            items.append((descriptor, warns, leaf.source_relpath))
        except Exception as exc:  # noqa: BLE001 - one bad leaf must not abort the sweep
            report["errors"].append({"leaf": leaf.source_relpath, "error": f"{type(exc).__name__}: {exc}"})

    seen_ids: dict[str, str] = {}
    for descriptor, warns, source_relpath in items:
        did = descriptor.id
        if did in seen_ids:
            report["errors"].append({"leaf": source_relpath, "error": f"id collision {did!r} (kept {seen_ids[did]})"})
            continue
        seen_ids[did] = source_relpath
        report["ids"][did] = source_relpath

        path = out_dir / f"{did}.yaml"
        if path.exists():
            try:
                old = DatasetDescriptor(**(yaml.safe_load(path.read_text(encoding="utf-8")) or {}))
            except Exception:  # noqa: BLE001 - unreadable existing descriptor: leave it for a human
                report["skipped"].append({"id": did, "reason": "existing descriptor unreadable"})
                continue
            if not (old.generation and old.generation.managed) and not force:
                report["skipped"].append({"id": did, "reason": "human-managed (generation.managed not set)"})
                continue
            if not force and metadata_hash(old) == metadata_hash(descriptor):
                report["skipped"].append({"id": did, "reason": "unchanged"})
                continue
            _write_descriptor(descriptor, path)
            report["updated"].append({"id": did, "warnings": warns})
        else:
            _write_descriptor(descriptor, path)
            report["created"].append({"id": did, "warnings": warns})

    # Reconciliation: prune managed orphans (descriptors the source no longer produces); never touch a
    # human-authored descriptor (only flag it). Always write a committed reconciliation report.
    new_ids = set(report["ids"])
    removed: list[str] = []
    kept_human: list[str] = []
    for path in sorted(out_dir.glob("*.yaml")):
        did = path.stem
        if did in new_ids:
            continue
        try:
            old = DatasetDescriptor(**(yaml.safe_load(path.read_text(encoding="utf-8")) or {}))
        except Exception:  # noqa: BLE001 - unreadable: leave it for a human
            continue
        if old.generation and old.generation.managed:
            removed.append(did)
            if prune:
                path.unlink()
                data_dir = Path(catalog_root) / "datasets" / did
                if data_dir.exists():
                    shutil.rmtree(data_dir)  # drop the orphan's tracked card/manifest/croissant + any local bytes
        else:
            kept_human.append(did)
    report["removed"] = sorted(removed)
    report["kept_human_orphans"] = sorted(kept_human)
    report["pruned"] = bool(prune)

    recon = {
        "source_root": str(source_root),
        "n_datasets": len(new_ids),
        "created": sorted(c["id"] for c in report["created"]),
        "updated": sorted(u["id"] for u in report["updated"]),
        "skipped": report["skipped"],
        "removed_managed_orphans": report["removed"],
        "pruned": bool(prune),
        "kept_human_orphans": report["kept_human_orphans"],
        "id_collisions": [e for e in report["errors"] if "collision" in e["error"]],
        "errors": [e for e in report["errors"] if "collision" not in e["error"]],
    }
    (Path(catalog_root) / "catalog" / "reconciliation.json").write_text(json.dumps(recon, indent=2, sort_keys=True), encoding="utf-8")
    return report
